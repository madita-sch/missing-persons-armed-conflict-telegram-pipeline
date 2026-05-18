# Import libraries
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction

# Create main function that builds the error analysis report with multiple sheets
# for different error types (classification, NER, translation, clustering,
# pseudonymization leaks) and a full debug sheet with all data side-by-side.
def build_error_analysis_report(df_pred, df_gold, output_path):
    from src.evaluation.evaluation_nlp import split_entities, is_match

    merged = df_pred.merge(df_gold, on="id", suffixes=("_pred", "_gold"))

    # Downstream filter: only rows where both gold and pred are missing-person cases
    downstream = merged[
        (merged["is_missing_gold"] == 1) &
        (merged["is_missing_pred"] == 1)
    ].copy()

    # Classification errors: all rows, no filter
    cls = merged[merged["is_missing_pred"] != merged["is_missing_gold"]][[
        "id",
        "is_missing_pred", "is_missing_gold",
        "text_clean_pred", "text_clean_gold",
        "text_clean_en_pred", "text_clean_en_gold",
    ]].copy()
    cls["error_type"] = cls.apply(
        lambda r: "FP_missing (pred=1, gold=0)" if r["is_missing_pred"] == 1
        else "FN_missing (pred=0, gold=1)", axis=1
    )

    # NER errors: gold=1 AND pred=1 only
    ner_rows = []
    ctx_cols = [
        "text_clean_pred", "text_clean_gold",
        "text_clean_en_pred", "text_clean_en_gold",
        "names_pred",    "names_gold",
        "location_pred", "location_gold",
        "dates_pred",    "dates_gold",
        "age_pred",      "age_gold",
        "names_en_pred", "names_en_gold",
        "location_en_pred", "location_en_gold",
        "dates_en_pred", "dates_en_gold",
    ]

    for _, row in downstream.iterrows():
        for col in ["names", "location", "dates", "age"]:
            pred = split_entities(row[f"{col}_pred"])
            gold = split_entities(row[f"{col}_gold"])

            for g in gold:
                if not any(is_match(g, p) for p in pred):
                    ner_rows.append({
                        "id": row["id"],
                        "error_type": "NER_FN",
                        "entity_type": col,
                        "missed_value (gold)": g,
                        "pred_had": "; ".join(pred) if pred else "",
                        **{c: row.get(c, "") for c in ctx_cols},
                    })

            for p in pred:
                if not any(is_match(p, g) for g in gold):
                    ner_rows.append({
                        "id": row["id"],
                        "error_type": "NER_FP",
                        "entity_type": col,
                        "missed_value (gold)": "",
                        "pred_had": p,
                        **{c: row.get(c, "") for c in ctx_cols},
                    })

    ner = pd.DataFrame(ner_rows)

    # Translation errors: gold=1 AND pred=1 only
    smoother = SmoothingFunction().method1
    trans_rows = []
    for _, row in downstream.iterrows():
        ref  = str(row.get("text_clean_en_gold", "")).split()
        pred = str(row.get("text_clean_en_pred", "")).split()
        if not ref or not pred:
            continue
        bleu = sentence_bleu([ref], pred, smoothing_function=smoother)
        trans_rows.append({
            "id":                   row["id"],
            "bleu_score":           round(bleu, 4),
            "bleu_pct":             f"{bleu*100:.1f}%",
            "quality_flag":         "Low" if bleu < 0.05 else "Med" if bleu < 0.30 else "OK",
            "text_clean":           row.get("text_clean_pred", ""),
            "text_clean_en_pred":   row.get("text_clean_en_pred", ""),
            "text_clean_en_gold":   row.get("text_clean_en_gold", ""),
        })
    translation = pd.DataFrame(trans_rows).sort_values("bleu_score")

    # Clustering errors: gold=1 AND pred=1 only
    cluster_rows = []
    clustered = downstream[downstream["cluster_id_gold"] != -1].copy()

    for gold_cid, group in clustered.groupby("cluster_id_gold"):
        pred_assignments = group["cluster_id_pred"].value_counts()
        if len(pred_assignments) == 1 and pred_assignments.index[0] != -1:
            continue  # correct clustering

        dominant_pred = pred_assignments.index[0]
        for _, row in group.iterrows():
            if row["cluster_id_pred"] != dominant_pred:
                error_type = "split_from_cluster" if row["cluster_id_pred"] != -1 else "missed_cluster"
                cluster_rows.append({
                    "id":                 row["id"],
                    "cluster_id_gold":    gold_cid,
                    "cluster_id_pred":    row["cluster_id_pred"],
                    "error_type":         error_type,
                    "names_pred":         row.get("names_pred", ""),
                    "names_gold":         row.get("names_gold", ""),
                    "location_pred":      row.get("location_pred", ""),
                    "text_clean_pred":    row.get("text_clean_pred", ""),
                    "text_clean_en_pred": row.get("text_clean_en_pred", ""),
                })

    for pred_cid, group in downstream[downstream["cluster_id_pred"] != -1].groupby("cluster_id_pred"):
        gold_assignments = group["cluster_id_gold"].value_counts()
        if len(gold_assignments) > 1:
            for _, row in group.iterrows():
                cluster_rows.append({
                    "id":                 row["id"],
                    "cluster_id_gold":    row["cluster_id_gold"],
                    "cluster_id_pred":    pred_cid,
                    "error_type":         "false_merge",
                    "names_pred":         row.get("names_pred", ""),
                    "names_gold":         row.get("names_gold", ""),
                    "location_pred":      row.get("location_pred", ""),
                    "text_clean_pred":    row.get("text_clean_pred", ""),
                    "text_clean_en_pred": row.get("text_clean_en_pred", ""),
                })

    clustering_errors = pd.DataFrame(cluster_rows).drop_duplicates(subset=["id", "error_type"])

    # Pseudonymization leaks: gold=1 AND pred=1 only
    phone_pattern = re.compile(r"\b\d{8,15}\b")
    leak_rows = []

    for _, row in downstream.iterrows():
        anon_text = str(row.get("text_clean_anon_pred", ""))
        leaks = []

        for name in str(row.get("names_pred", "")).split(";"):
            name = name.strip()
            if name and len(name) > 2 and name in anon_text:
                leaks.append(("name_leak", name))

        for phone in phone_pattern.findall(anon_text):
            leaks.append(("phone_leak", phone))

        if leaks:
            name_leaks  = [v for t, v in leaks if t == "name_leak"]
            phone_leaks = [v for t, v in leaks if t == "phone_leak"]
            leak_rows.append({
                "id":                   row["id"],
                "leaked_names":         "; ".join(name_leaks)  or "",
                "leaked_phones":        "; ".join(phone_leaks) or "",
                "n_name_leaks":         len(name_leaks),
                "n_phone_leaks":        len(phone_leaks),
                "names_pred":           row.get("names_pred", ""),
                "text_clean_pred":      row.get("text_clean_pred", ""),
                "text_clean_anon_pred": anon_text,
                "text_clean_anon_gold": row.get("text_clean_anon_gold", ""),
            })

    leak = pd.DataFrame(leak_rows)

    # Full debug sheet: gold=1 AND pred=1 only
    full_cols = [
        "id",
        "is_missing_pred", "is_missing_gold",
        "cluster_id_pred", "cluster_id_gold",
        "names_pred",    "names_gold",
        "location_pred", "location_gold",
        "dates_pred",    "dates_gold",
        "age_pred",      "age_gold",
        "text_clean_pred", "text_clean_gold",
        "text_clean_en_pred", "text_clean_en_gold",
    ]
    full = downstream[[c for c in full_cols if c in downstream.columns]]

    print(f"  Classification errors : {len(cls)} rows")
    print(f"  NER errors            : {len(ner)} rows")
    print(f"  Translation           : {len(translation)} rows")
    print(f"  Clustering errors     : {len(clustering_errors)} rows")
    print(f"  Pseudonymization      : {len(leak)} rows")
    print(f"  Full debug            : {len(full)} rows")

    # Write all sheets to Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        cls.to_excel(writer,               sheet_name="classification",    index=False)
        ner.to_excel(writer,               sheet_name="ner_errors",        index=False)
        translation.to_excel(writer,       sheet_name="translation",       index=False)
        clustering_errors.to_excel(writer, sheet_name="clustering_errors", index=False)
        leak.to_excel(writer,              sheet_name="pseudonymization",  index=False)
        full.to_excel(writer,              sheet_name="full_debug",        index=False)

    # Apply formatting: bold headers, freeze top row, wrap text, set column widths
    _format_excel(output_path)
    print(f"Saved -> {output_path}")

    return {
        "classification":    cls,
        "ner":               ner,
        "translation":       translation,
        "clustering_errors": clustering_errors,
        "pseudonymization":  leak,
        "full":              full,
    }


def _format_excel(path):
    # formatting: bold header row, freeze pane, wrap text, column widths
    wb = load_workbook(path)

    # Column widths per sheet
    widths = {
        "classification":   {"A": 8,  "B": 14, "C": 14, "D": 50, "E": 50, "F": 50, "G": 50, "H": 28},
        "ner_errors":       {"A": 8,  "B": 14, "C": 14, "D": 30, "E": 30, "F": 45, "G": 45,
                             "H": 45, "I": 45, "J": 28, "K": 28, "L": 28, "M": 28,
                             "N": 18, "O": 18, "P": 10, "Q": 10, "R": 28, "S": 28,
                             "T": 28, "U": 28, "V": 18, "W": 18},
        "translation":      {"A": 8,  "B": 10, "C": 10, "D": 12, "E": 50, "F": 50, "G": 50},
        "clustering_errors":{"A": 8,  "B": 14, "C": 14, "D": 18, "E": 28, "F": 28,
                             "G": 28, "H": 50, "I": 50},
        "pseudonymization": {"A": 8,  "B": 35, "C": 25, "D": 12, "E": 12,
                             "F": 30, "G": 50, "H": 50, "I": 50},
        "full_debug":       {"A": 8,  "B": 14, "C": 14, "D": 14, "E": 14,
                             "F": 28, "G": 28, "H": 28, "I": 28, "J": 18,
                             "K": 18, "L": 10, "M": 10, "N": 50, "O": 50, "P": 50, "Q": 50},
    }

    header_font = Font(bold=True, name="Arial", size=10)
    body_font   = Font(name="Arial", size=9)
    wrap        = Alignment(wrap_text=True, vertical="top")
    header_fill = PatternFill("solid", fgColor="D9D9D9")  # light grey header

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"

        # Bold grey header row
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 25

        # Wrap text on all data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font      = body_font
                cell.alignment = wrap
            ws.row_dimensions[row[0].row].height = 55

        # Set column widths
        for col, w in widths.get(sheet_name, {}).items():
            ws.column_dimensions[col].width = w

    wb.save(path)
